# Plot or export cal/phy partition groupings and/or detector usability.
#
# Example usage:
#   # all cal groupings (escale-combined), both xlsx and pdf
#   python plot_partition_groupings.py --type cal --grouping cal --output both
#
#   # napoli26 key, escale-specific groupings, xlsx only
#   python plot_partition_groupings.py --type cal --key napoli26 --grouping escale --output xlsx
#
#   # psd groupings, pdf only
#   python plot_partition_groupings.py --type cal --key napoli26 --grouping psd --output pdf
#
#   # usability only (no grouping needed), pdf
#   python plot_partition_groupings.py --type cal --key napoli26 --output pdf
#
#   # phy groupings
#   python plot_partition_groupings.py --type phy --grouping phy --output both

from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from legendmeta import LegendMetadata

GROUPING_YAML_MAP = {
    "cal": "cal_groupings.yaml",
    "phy": "phy_groupings.yaml",
    "escale": "escale_groupings.yaml",  # combined og + escale
    "psd": "psd_groupings.yaml",  # combined og + psd
}

PART_COLOURS = [
    "BDD7EE",
    "9DC3E6",
    "C6EFCE",
    "A9D18E",
    "FFE699",
    "FFD966",
    "FCE4CC",
    "F4B183",
    "E2D3F0",
    "C4A8E0",
    "C7EAE4",
    "8FD0C8",
    "FADADD",
    "F4A7B0",
    "E9F5C2",
    "D4ED8A",
    "F9D6E8",
    "F0A8CC",
    "D6DCE4",
    "ACB9CA",
]
STRING_CELL_SHADES = ["F2F2F2", "E4E4E4"]
PERIOD_COLOURS = [
    "C0D7EE",
    "C6DEBB",
    "FFF2A0",
    "F5CBCC",
    "DDD0EA",
    "FDDCB5",
    "C7EAE4",
    "E8E8E8",
    "D9EAD3",
    "FDE8C8",
]
EMPTY_COLOR = "FFFFFF"
ON_COLOR = "6BAE75"
AC_COLOR = "FFA500"
OFF_COLOR = "FF4444"


def load_yaml(path: str) -> dict:
    with Path.open(path) as f:
        return yaml.safe_load(f)


def expand_run_list(value: list | str) -> list[str]:
    """Expand a YAML run value to a flat list of run strings."""
    if isinstance(value, list):
        result = []
        for item in value:
            s = str(item)
            if ".." in s:
                start, end = s.split("..")
                result.extend(
                    [f"r{n:03d}" for n in range(int(start[1:]), int(end[1:]) + 1)]
                )
            else:
                result.append(s)
        return result
    s = str(value)
    if ".." in s:
        start, end = s.split("..")
        return [f"r{n:03d}" for n in range(int(start[1:]), int(end[1:]) + 1)]
    return [s]


def compress_runs(runs: set[str] | list[str]) -> str | list[str]:
    """Compact a collection of run strings.
    Fully consecutive  → single range string  e.g. 'r000..r006'
    Non-consecutive    → flat list of individual runs  e.g. ['r000', 'r002', 'r003']
    Never mix ranges and singles inside a list."""
    nums = sorted(int(r[1:]) for r in runs)
    if nums == list(range(nums[0], nums[-1] + 1)):
        if nums[0] == nums[-1]:
            return f"r{nums[0]:03d}"
        return f"r{nums[0]:03d}..r{nums[-1]:03d}"
    return [f"r{n:03d}" for n in nums]


def partition_num(part_name: str) -> int:
    """calgroup008b  →  8"""
    return int(part_name.split("group")[-1][:-1])


def partition_sub(part_name: str) -> str:
    """calgroup008b  →  'b'"""
    return part_name.split("group")[-1][-1]


def partition_label(part_name: str) -> str:
    """calgroup008b  →  '8b'  (short display label)"""
    temp = part_name.split("group")[-1]
    return f"{int(temp[:-1])}{temp[-1]}"


def partition_base(part_name: str) -> str:
    """calgroup008b  →  '008'  (zero-padded base number, for sub-partition grouping)"""
    return part_name.split("group")[-1][:-1]


def col_sort_key(period_run: tuple) -> tuple:
    """Sort key for (period, run) tuples e.g. ('p18', 'r003') → (18, 3)"""
    period, run = period_run
    return (int(period[1:]), int(run[1:]))


def hex_to_rgb01(h: str) -> tuple:
    """'FFA500' → (1.0, 0.647, 0.0)"""
    h = h.lstrip("#")
    return tuple(int(h[i : i + 2], 16) / 255 for i in (0, 2, 4))


def build_period_run_map(
    partition_dict: dict, min_part: int, skip_single: bool = False
) -> dict:
    """Flatten a default-block dict into {(period, run): partition_name}.
    e.g. {('p16', 'r000'): 'calgroup008a', ('p16', 'r001'): 'calgroup008a', ...}"""
    result = {}
    for part, period_runs in partition_dict.items():
        if partition_num(part) < min_part:
            continue
        if not isinstance(period_runs, dict):
            continue
        for period, runs in period_runs.items():
            expanded_runs = expand_run_list(runs)
            if not runs:
                continue
            if skip_single and len(expanded_runs) == 1:
                continue
            for run in expanded_runs:
                result[(period, run)] = part
    return result


def merge_with_defaults(
    hpge_partitions: dict, default_partitions: dict, min_part: int
) -> dict:
    """Merge a detector's partition overrides with the defaults.
    Returns {(period, run): partition_name} for all valid runs."""
    all_partitions = {
        part
        for part in list(hpge_partitions) + list(default_partitions)
        if partition_num(part) >= min_part
    }
    result = {}
    for part in all_partitions:
        default_source = default_partitions.get(part, {})
        if not isinstance(default_source, dict):
            default_source = {}

        if part in hpge_partitions:
            device_source = hpge_partitions[part]
            if not isinstance(device_source, dict):
                continue
            for period, runs in device_source.items():
                expanded_runs = expand_run_list(runs)
                if len(expanded_runs) == 1:
                    continue  # single-run device override → leave blank
                for run in expanded_runs:
                    result[(period, run)] = part
            # periods in default but absent in device override → inherit default
            for period, runs in default_source.items():
                if period not in device_source:
                    for run in expand_run_list(runs):
                        result[(period, run)] = part
        else:
            for period, runs in default_source.items():
                for run in expand_run_list(runs):
                    result[(period, run)] = part
    return result


def xl_fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)


parser = argparse.ArgumentParser(description="Plot cal or phy partition groupings")
parser.add_argument(
    "--type",
    choices=["cal", "phy"],
    default="cal",
    help="cal or phy grouping (used for runlist lookup)",
)
parser.add_argument("--key", default=None, help="runlist key e.g. napoli26")
parser.add_argument(
    "--grouping",
    choices=list(GROUPING_YAML_MAP),
    default=None,
    help="which grouping yaml: cal, phy, escale, psd  (omit for usability-only)",
)
parser.add_argument(
    "--usability",
    action="store_true",
    help="show usability fills only even when a grouping is provided",
)
parser.add_argument(
    "--output",
    choices=["xlsx", "pdf", "both"],
    default="both",
    help="output format(s) to produce",
)
args = parser.parse_args()

# if no grouping is given we are implicitly in usability-only mode
USAB_ONLY = args.grouping is None or args.usability

RUNLIST_YAML = "runlists.yaml"
RUNINFO_YAML = "runinfo.yaml"

if USAB_ONLY:
    _base = f"{args.type}_usability"
elif args.type == "phy":
    _base = "phy_groupings"
else:
    _base = f"{args.grouping}_cal_groupings"
if args.key:
    _base = f"{args.key}_{_base}"
XL_OUT = f"{_base}_display.xlsx"
PDF_OUT = f"{_base}_display.pdf"

MIN_PART = 0 if args.key else 8
FONT = "Arial"


# ── colours, fonts and borders ────────────────────────────────────────────────

THICK = Side(style="medium", color="333333")
NORMAL = Side(style="thin", color="888888")
HAIR = Side(style="hair", color="CCCCCC")

# ── load & build mappings/dictionaries ────────────────────────────────────────

runlists = load_yaml(RUNLIST_YAML)
runinfo = load_yaml(RUNINFO_YAML)

meta = LegendMetadata()
chmap = meta.channelmap("20250828T033011Z")
str_pos = {}
for ged, item in chmap.items():
    if item["system"] != "geds":
        continue
    str_pos[ged] = {
        "string": item["location"]["string"],
        "position": item["location"]["position"],
    }

if args.grouping:
    groupings = load_yaml(GROUPING_YAML_MAP[args.grouping])
    defaults = groupings["default"]
    # default_map: {('p16', 'r000'): 'calgroup008a', ('p16', 'r001'): 'calgroup008a', ...}
    default_map = build_period_run_map(defaults, skip_single=False)
    hpge_maps = {}
    for hpge in str_pos:
        if hpge in groupings:
            hpge_maps[hpge] = merge_with_defaults(groupings[hpge], defaults)
        else:
            hpge_maps[hpge] = default_map
    # hpge_maps: {'V02160A': {('p16', 'r000'): 'calgroup008a', ('p18', 'r002'): 'calgroup009a', ...}, ...}
    # all_short_labels: ['8a', '8b', '9a', '10a', ...]
    all_short_labels = sorted(
        {partition_label(part) for m in hpge_maps.values() for part in m.values()}
    )
    # label_colour_map: {'8a': 'BDD7EE', '8b': '9DC3E6', '9a': 'C6EFCE', ...}
    label_colour_map = {
        lbl: PART_COLOURS[i % len(PART_COLOURS)]
        for i, lbl in enumerate(all_short_labels)
    }
else:
    default_map = {}
    hpge_maps = {hpge: {} for hpge in str_pos}
    label_colour_map = {}

hpges = sorted(
    str_pos.keys(), key=lambda d: (str_pos[d]["string"], str_pos[d]["position"])
)

string_groups = {}
for hpge in hpges:
    string = str_pos[hpge]["string"]
    if string not in string_groups:
        string_groups[string] = []
    string_groups[string].append(hpge)

# string_groups: {1: ['V02160A', 'V02162B', ...], 2: ['B00032C', ...], ...}
strings = sorted(string_groups.keys())
# string_shade_map: {1: 'F2F2F2', 2: 'E4E4E4', 3: 'F2F2F2', ...}
string_shade_map = {s: STRING_CELL_SHADES[i % 2] for i, s in enumerate(strings)}

# ── columns ───────────────────────────────────────────────────────────────────

output = {}
if args.key:
    key_data = runlists[args.key][args.type]
    for period, runs_spec in key_data.items():
        runs = expand_run_list(runs_spec)
        if runs:
            max_run = max(int(r[1:]) for r in runs)
            if period not in output:
                output[period] = {"max": 0}
            output[period]["max"] = max(output[period]["max"], max_run)
elif args.grouping:
    for tup, _name in default_map.items():
        period, run = tup[0], tup[1]
        if period not in output:
            output[period] = {"min": 0, "max": 0}
        run_num = int(run[1:])
        output[period]["max"] = max(output[period]["max"], run_num)
else:
    msg = "--key is required when no --grouping is specified"
    raise ValueError(msg)

# sorted_cols: [('p16', 'r000'), ('p16', 'r001'), ..., ('p18', 'r000'), ...]
all_period_runs = [
    (period, f"r{run:03d}")
    for period, info in output.items()
    for run in range(info["max"] + 1)
]
sorted_cols = sorted(all_period_runs, key=col_sort_key)

# periods: ['p16', 'p18', 'p19']
periods = sorted({period for period, run in sorted_cols}, key=lambda p: int(p[1:]))

# period_groups: {'p16': [('p16','r000'), ('p16','r001'), ...], 'p18': [...], ...}
period_groups = {}
for period_run in sorted_cols:
    period, run = period_run[0], period_run[1]
    if period not in period_groups:
        period_groups[period] = []
    period_groups[period].append(period_run)

# ── usability lookup ──────────────────────────────────────────────────────────
# usab_map: {('p16', 'r000', 'V02160A'): 'on', ('p16', 'r000', 'V05268A'): 'off', ...}

usab_map = {}
for period_run in sorted_cols:
    period, run = period_run[0], period_run[1]
    timestamp = runinfo[period][run][args.type]["start_key"]
    run_chmap = meta.channelmap(timestamp)
    for hpge, item in run_chmap.items():
        if item["system"] != "geds":
            continue
        usab_map[(period, run, hpge)] = item["analysis"]["usability"]

# ── cell colour helper ────────────────────────────────────────────────────────


def cell_colours(
    hpge: str, period: str, run: str, part_map: dict[tuple[str, str], str]
) -> tuple[str, str]:
    """Return (fill_hex, label) for one cell."""
    part = part_map.get((period, run))
    status = usab_map.get((period, run, hpge))

    if USAB_ONLY:
        # usability-only: colour by status, no partition labels
        if status == "off":
            return OFF_COLOR, ""
        if status == "ac":
            return AC_COLOR, ""
        if status == "on":
            return ON_COLOR, ""
        return EMPTY_COLOR, ""
    # partition mode: show grouping label, usability overrides fill
    if part:
        lbl = partition_label(part)
        base_hex = label_colour_map.get(lbl, "CCCCCC")
    else:
        base_hex, lbl = EMPTY_COLOR, ""
    if status == "off":
        return OFF_COLOR, lbl
    if status == "ac":
        return AC_COLOR, lbl
    return base_hex, lbl


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════

if args.output in ("xlsx", "both"):
    COL_OFFSET = 3
    GRP_ROW = 1
    HEADER_ROW = 2
    DATA_START = 3

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Groupings"

    # period group header
    ws.row_dimensions[GRP_ROW].height = 18
    col_cursor = COL_OFFSET
    for period in periods:
        runs = period_groups.get(period, [])
        if not runs:
            continue
        sc, ec = col_cursor, col_cursor + len(runs) - 1
        sl, el = get_column_letter(sc), get_column_letter(ec)
        ws.merge_cells(f"{sl}{GRP_ROW}:{el}{GRP_ROW}")
        c = ws[f"{sl}{GRP_ROW}"]
        c.value = f"{period}  ({runs[0][1]} - {runs[-1][1]})"
        c.fill = xl_fill(PERIOD_COLOURS[periods.index(period)])
        c.font = Font(bold=True, name=FONT, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        col_cursor += len(runs)

    # rotated header row
    ws.row_dimensions[HEADER_ROW].height = 70
    hdr_dark = xl_fill("2F4F7F")
    for col, label in enumerate(
        ["String", "Detector"] + [f"{p}_{r}" for p, r in sorted_cols], 1
    ):
        c = ws.cell(HEADER_ROW, col, label)
        c.fill = hdr_dark
        c.font = Font(bold=True, name=FONT, size=9, color="FFFFFF")
        c.alignment = (
            Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col <= 2
            else Alignment(horizontal="center", vertical="bottom", text_rotation=90)
        )

    # first-column positions for period group borders
    p_first_col = {}
    col_cursor = COL_OFFSET
    for period in periods:
        if period_groups.get(period):
            p_first_col[period] = col_cursor
            col_cursor += len(period_groups[period])
    last_col = COL_OFFSET + len(sorted_cols) - 1

    current_row = DATA_START
    for string_num in strings:
        string_hpges = string_groups[string_num]
        bg = string_shade_map[string_num]
        string_start = current_row

        for hpge_idx, hpge in enumerate(string_hpges):
            row = current_row
            part_map = hpge_maps[hpge]
            is_first = hpge_idx == 0
            is_last = hpge_idx == len(string_hpges) - 1
            top = THICK if is_first else NORMAL
            bottom = THICK if is_last else None

            ca = ws.cell(row, 1)
            ca.fill = xl_fill(bg)
            ca.border = Border(
                left=THICK,
                right=NORMAL,
                top=THICK if is_first else None,
                bottom=THICK if is_last else None,
            )

            cb = ws.cell(row, 2, hpge)
            cb.fill = xl_fill(bg)
            cb.font = Font(bold=True, name=FONT, size=9)
            cb.alignment = Alignment(horizontal="center", vertical="center")
            cb.border = Border(left=NORMAL, right=THICK, top=top, bottom=bottom)

            for run_idx, (period, run) in enumerate(sorted_cols):
                col = COL_OFFSET + run_idx
                cell = ws.cell(row, col)
                cell.border = Border(
                    left=THICK if col in p_first_col.values() else HAIR,
                    right=THICK if col == last_col else None,
                    top=top,
                    bottom=bottom,
                )
                fill_hex, lbl = cell_colours(hpge, period, run, part_map)
                cell.fill = xl_fill(fill_hex)
                if lbl:
                    cell.value = lbl
                    cell.font = Font(name=FONT, size=8)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        string_end = current_row - 1
        ws.merge_cells(f"A{string_start}:A{string_end}")
        c = ws.cell(string_start, 1)
        c.value = f"String {string_num}"
        c.font = Font(bold=True, name=FONT, size=10)
        c.fill = xl_fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 10
    for i in range(len(sorted_cols)):
        ws.column_dimensions[get_column_letter(COL_OFFSET + i)].width = 3.8
    for row in range(DATA_START, current_row):
        ws.row_dimensions[row].height = 15
    ws.freeze_panes = f"{get_column_letter(COL_OFFSET)}{DATA_START}"

    wb.save(XL_OUT)
    # print(
    #    f"Excel saved: {XL_OUT}  ({len(hpges)} detectors, {len(sorted_cols)} columns)"
    # )

# ══════════════════════════════════════════════════════════════════════════════
# PLOT pdf
# ══════════════════════════════════════════════════════════════════════════════

if args.output in ("pdf", "both"):
    n_det = len(hpges)
    n_cols = len(sorted_cols)

    label_grid = [[""] * n_cols for _ in range(n_det)]
    color_grid = np.ones((n_det, n_cols, 4))
    for row_i, det in enumerate(hpges):
        part_map = hpge_maps[det]
        for col_i, (period, run) in enumerate(sorted_cols):
            fill_hex, lbl = cell_colours(det, period, run, part_map)
            color_grid[row_i, col_i] = (*hex_to_rgb01(fill_hex), 1.0)
            label_grid[row_i][col_i] = lbl

    # string boundary rows and start positions
    row_cursor = 0
    string_row_start = {}
    string_boundaries = []
    for string_num in strings:
        string_row_start[string_num] = row_cursor
        row_cursor += len(string_groups[string_num])
        string_boundaries.append(row_cursor)
    string_boundaries = string_boundaries[:-1]

    # period boundary columns and start positions
    col_cursor = 0
    period_col_start = {}
    period_boundaries = []
    for period in periods:
        runs = period_groups.get(period, [])
        if not runs:
            continue
        period_col_start[period] = col_cursor
        col_cursor += len(runs)
        period_boundaries.append(col_cursor)
    period_boundaries = period_boundaries[:-1]

    STR_W = 2.0
    DET_W = 3.5
    RUN_LBL_H = 1.0
    PER_BAR_H = 0.8
    HDR_H = RUN_LBL_H + PER_BAR_H + 0.2

    x_str0 = -(STR_W + DET_W)
    x_det0 = -DET_W
    x_full = x_str0

    cell_in = 0.28
    fig_w = (STR_W + DET_W + n_cols) * cell_in + 0.2
    fig_h = (HDR_H + n_det) * cell_in + 0.3

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(x_full - 0.05, n_cols + 0.05)
    ax.set_ylim(n_det, -HDR_H)
    ax.axis("off")

    # row backgrounds
    for string_num in strings:
        start = string_row_start[string_num]
        size = len(string_groups[string_num])
        ax.add_patch(
            mpatches.Rectangle(
                (x_full, start),
                n_cols - x_full,
                size,
                facecolor=hex_to_rgb01(string_shade_map[string_num]),
                edgecolor="none",
                zorder=0,
            )
        )

    # run cells
    for row_i in range(n_det):
        for col_i in range(n_cols):
            ax.add_patch(
                mpatches.Rectangle(
                    (col_i, row_i),
                    1,
                    1,
                    facecolor=color_grid[row_i, col_i],
                    edgecolor="#CCCCCC",
                    linewidth=0.3,
                    zorder=1,
                )
            )
            lbl = label_grid[row_i][col_i]
            if lbl:
                ax.text(
                    col_i + 0.5,
                    row_i + 0.5,
                    lbl,
                    ha="center",
                    va="center",
                    fontsize=5.5,
                    color="#333333",
                    zorder=2,
                )

    # detector name column
    for row_i, det in enumerate(hpges):
        ax.add_patch(
            mpatches.Rectangle(
                (x_det0, row_i),
                DET_W,
                1,
                facecolor="none",
                edgecolor="#AAAAAA",
                linewidth=0.3,
                zorder=1,
            )
        )
        ax.text(
            x_det0 + DET_W / 2,
            row_i + 0.5,
            det,
            ha="center",
            va="center",
            fontsize=6,
            fontweight="bold",
            fontfamily="monospace",
            zorder=2,
        )

    # string label column
    for string_num in strings:
        start = string_row_start[string_num]
        size = len(string_groups[string_num])
        ax.add_patch(
            mpatches.Rectangle(
                (x_str0, start),
                STR_W,
                size,
                facecolor="none",
                edgecolor="#555555",
                linewidth=0.8,
                zorder=3,
            )
        )
        ax.text(
            x_str0 + STR_W / 2,
            start + size / 2,
            f"Str\n{string_num}",
            ha="center",
            va="center",
            fontsize=6.5,
            fontweight="bold",
            zorder=4,
        )

    # string + period boundary lines
    for sb in string_boundaries:
        ax.plot([x_full, n_cols], [sb, sb], color="#333333", linewidth=1.2, zorder=5)
    for pb in period_boundaries:
        ax.plot([pb, pb], [0, n_det], color="#333333", linewidth=1.2, zorder=5)

    # rotated run labels
    bar_y0 = -(RUN_LBL_H + PER_BAR_H)
    bar_y_mid = bar_y0 + PER_BAR_H / 2
    for col_i, (_period, run) in enumerate(sorted_cols):
        ax.text(
            col_i + 0.5,
            0,
            run,
            ha="center",
            va="bottom",
            fontsize=5.5,
            rotation=90,
            color="#222222",
            zorder=4,
        )

    # period header bars
    for period in periods:
        runs = period_groups.get(period, [])
        if not runs:
            continue
        sc = period_col_start[period]
        ax.add_patch(
            mpatches.Rectangle(
                (sc, bar_y0),
                len(runs),
                PER_BAR_H,
                facecolor=hex_to_rgb01(PERIOD_COLOURS[periods.index(period)]),
                edgecolor="#555555",
                linewidth=0.8,
                zorder=3,
            )
        )
        ax.text(
            sc + len(runs) / 2,
            bar_y_mid,
            f"{period}  ({runs[0][1]}-{runs[-1][1]})",
            ha="center",
            va="center",
            fontsize=7,
            fontweight="bold",
            zorder=4,
        )

    # String/Detector column headers
    for x0, w, label in [(x_str0, STR_W, "String"), (x_det0, DET_W, "Detector")]:
        ax.add_patch(
            mpatches.Rectangle(
                (x0, bar_y0),
                w,
                PER_BAR_H,
                facecolor=hex_to_rgb01("2F4F7F"),
                edgecolor="#333333",
                linewidth=0.8,
                zorder=3,
            )
        )
        ax.text(
            x0 + w / 2,
            bar_y_mid,
            label,
            ha="center",
            va="center",
            fontsize=7,
            fontweight="bold",
            color="white",
            zorder=4,
        )

    plt.tight_layout(pad=0.2)
    fig.savefig(PDF_OUT, bbox_inches="tight")
    # print(f"Plot saved:  {PDF_OUT}")
