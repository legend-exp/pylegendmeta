from __future__ import annotations

import re

import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from dbetto import Props
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from legendmeta import LegendMetadata

PART_CMAP = "tab20"
PERIOD_CMAP = "Pastel2"
STRING_CELL_SHADES = ["F2F2F2", "E4E4E4"]
EMPTY_COLOR = "FFFFFF"
ON_COLOR = "6BAE75"
AC_COLOR = "FFA500"
OFF_COLOR = "FF4444"

THICK = Side(style="medium", color="333333")
NORMAL = Side(style="thin", color="888888")
HAIR = Side(style="hair", color="CCCCCC")
FONT = "Arial"

GROUPING_YAML_MAP: dict[str, str] = {
    "cal": "dataprod/config/partitions/cal_groupings.yaml",
    "phy": "dataprod/config/partitions/phy_groupings.yaml",
    "escale": "dataprod/config/partitions/escale_groupings.yaml",
    "psd": "dataprod/config/partitions/psd_groupings.yaml",
}


def expand_run_list(value: list | str) -> list[str]:
    """Expand a YAML run value to a flat list of run strings."""
    if isinstance(value, list):
        result = []
        for item in value:
            s = str(item)
            if ".." in s:
                start, end = s.split("..", 1)
                result.extend(
                    [f"r{n:03d}" for n in range(int(start[1:]), int(end[1:]) + 1)]
                )
            else:
                result.append(s)
        return result
    s = str(value)
    if ".." in s:
        start, end = s.split("..", 1)
        return [f"r{n:03d}" for n in range(int(start[1:]), int(end[1:]) + 1)]
    return [s]


def partition_num(part_name: str) -> int:
    """calgroup008b  →  8"""
    return int(part_name.split("group")[-1][:-1])


def partition_label(part_name: str) -> str:
    """calgroup008b  →  '8b'  (short display label)"""
    temp = part_name.split("group")[-1]
    return f"{int(temp[:-1])}{temp[-1]}"


def col_sort_key(period_run: tuple) -> tuple:
    """Sort key for (period, run) tuples e.g. ('p18', 'r003') → (18, 3)"""
    period, run = period_run
    return (int(period[1:]), int(run[1:]))


def hex_to_rgb01(h: str) -> tuple:
    """'FFA500' → (1.0, 0.647, 0.0)"""
    h = h.lstrip("#")
    return tuple(int(h[i : i + 2], 16) / 255 for i in (0, 2, 4))


def build_period_run_map(
    partition_dict: dict, min_part: int, skip_single: bool = False
) -> dict:
    """Flatten a default-block dict into {(period, run): partition_name}."""
    result = {}
    for part, period_runs in partition_dict.items():
        if partition_num(part) < min_part or not isinstance(period_runs, dict):
            continue
        for period, runs in period_runs.items():
            expanded = expand_run_list(runs)
            if not expanded or (skip_single and len(expanded) == 1):
                continue
            for run in expanded:
                result[(period, run)] = part
    return result


def merge_with_defaults(
    hpge_partitions: dict, default_partitions: dict, min_part: int
) -> dict:
    """Merge a detector's partition overrides with the defaults.
    Returns {(period, run): partition_name} for all valid runs."""
    all_partitions = {
        part
        for part in list(hpge_partitions) + list(default_partitions)
        if partition_num(part) >= min_part
    }
    result = {}
    for part in all_partitions:
        default_source = default_partitions.get(part, {})
        if not isinstance(default_source, dict):
            default_source = {}
        if part in hpge_partitions:
            device_source = hpge_partitions[part]
            if not isinstance(device_source, dict):
                continue
            for period, runs in device_source.items():
                expanded = expand_run_list(runs)
                if len(expanded) == 1:
                    continue  # single-run device override → leave blank
                for run in expanded:
                    result[(period, run)] = part
            for period, runs in default_source.items():
                if period not in device_source:
                    for run in expand_run_list(runs):
                        result[(period, run)] = part
        else:
            for period, runs in default_source.items():
                for run in expand_run_list(runs):
                    result[(period, run)] = part
    return result


def cmap_hex(cmap_name: str, n: int) -> list[str]:
    """Sample *n* evenly-spaced colours from a matplotlib colormap, returning hex strings."""
    cmap = plt.get_cmap(cmap_name)
    return [
        f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"
        for r, g, b, _ in (cmap(i / max(n - 1, 1)) for i in range(n))
    ]


def xl_fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)


def _build_run_layout(period: str, run: str, type: str) -> dict:
    """Load metadata for a single run.

    Returns a dict with keys: str_pos, usab_map, psd_map, period, run.
    """
    runinfo = Props.read_from("runinfo.yaml")
    timestamp = runinfo[period][run][type]["start_key"]

    meta = LegendMetadata()
    run_chmap = meta.channelmap(timestamp)

    str_pos = {}
    usab_map = {}
    psd_map = {}
    for hpge, item in run_chmap.items():
        if item["system"] != "geds":
            continue
        str_pos[hpge] = {
            "string": item["location"]["string"],
            "position": item["location"]["position"],
        }
        analysis = item.get("analysis")
        if not analysis:
            continue
        usability = analysis.get("usability")
        if usability is not None:
            usab_map[hpge] = usability
        if "psd" in analysis:
            psd_map[hpge] = analysis["psd"]

    return {
        "str_pos": str_pos,
        "usab_map": usab_map,
        "psd_map": psd_map,
        "period": period,
        "run": run,
    }


def _render_run(
    layout: dict, hpge_maps: dict, cell_colours, output: str | None
) -> None:
    """Render a single-run 2D spatial array (strings x positions).

    ``cell_colours(hpge, part_map) -> (fill_hex, label)``
    """
    str_pos = layout["str_pos"]

    # {string: {position: hpge}}
    grid: dict[int, dict[int, str]] = {}
    for hpge, loc in str_pos.items():
        grid.setdefault(loc["string"], {})[loc["position"]] = hpge

    strings = sorted(grid)
    max_pos = max((p for s in grid.values() for p in s), default=0)
    if not grid or not max_pos:
        return

    if output is not None and not output.endswith((".xlsx", ".pdf")):
        msg = f"Unsupported output format: {output!r}. Use '.xlsx' or '.pdf'."
        raise ValueError(msg)

    if output is not None and output.endswith(".xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Array"

        # header row: string numbers
        for col_i, string_num in enumerate(strings, 2):
            c = ws.cell(1, col_i, f"Str {string_num}")
            c.fill = xl_fill("2F4F7F")
            c.font = Font(bold=True, name=FONT, size=9, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")

        # header column: position numbers
        for pos in range(1, max_pos + 1):
            c = ws.cell(pos + 1, 1, pos)
            c.font = Font(bold=True, name=FONT, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")

        for col_i, string_num in enumerate(strings, 2):
            for pos in range(1, max_pos + 1):
                hpge = grid[string_num].get(pos)
                row = pos + 1
                cell = ws.cell(row, col_i)
                if hpge:
                    fill_hex, lbl = cell_colours(hpge, hpge_maps.get(hpge, {}))
                    cell.fill = xl_fill(fill_hex)
                    cell.value = f"{hpge}\n{lbl}" if lbl else hpge
                    cell.font = Font(name=FONT, size=8)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                else:
                    cell.fill = xl_fill("EEEEEE")

        for col_i in range(1, len(strings) + 2):
            ws.column_dimensions[get_column_letter(col_i)].width = 11
        for row in range(1, max_pos + 2):
            ws.row_dimensions[row].height = 30

        wb.save(output)

    if output is None or output.endswith(".pdf"):
        n_str = len(strings)
        cell_w = 1.0
        fig, ax = plt.subplots(figsize=(n_str * cell_w + 0.5, max_pos * cell_w + 0.5))
        ax.set_xlim(-0.05, n_str + 0.05)
        ax.set_ylim(max_pos + 0.05, -0.05)
        ax.axis("off")

        for col_i, string_num in enumerate(strings):
            ax.text(
                col_i + 0.5,
                -0.3,
                f"Str {string_num}",
                ha="center",
                va="bottom",
                fontsize=7,
                fontweight="bold",
            )
            for pos in range(1, max_pos + 1):
                hpge = grid[string_num].get(pos)
                row_i = pos - 1
                if hpge:
                    fill_hex, lbl = cell_colours(hpge, hpge_maps.get(hpge, {}))
                    ax.add_patch(
                        mpatches.Rectangle(
                            (col_i, row_i),
                            1,
                            1,
                            facecolor=hex_to_rgb01(fill_hex),
                            edgecolor="#888888",
                            linewidth=0.5,
                        )
                    )
                    ax.text(
                        col_i + 0.5,
                        row_i + 0.35,
                        hpge,
                        ha="center",
                        va="center",
                        fontsize=4.5,
                        fontfamily="monospace",
                        fontweight="bold",
                    )
                    if lbl:
                        ax.text(
                            col_i + 0.5,
                            row_i + 0.65,
                            lbl,
                            ha="center",
                            va="center",
                            fontsize=5,
                            color="#333333",
                        )
                else:
                    ax.add_patch(
                        mpatches.Rectangle(
                            (col_i, row_i),
                            1,
                            1,
                            facecolor=hex_to_rgb01("EEEEEE"),
                            edgecolor="#CCCCCC",
                            linewidth=0.3,
                        )
                    )

        period = layout.get("period", "")
        run = layout.get("run", "")
        if period and run:
            ax.set_title(f"{period} {run}", fontsize=9, fontweight="bold")

        plt.tight_layout(pad=0.3)
        if output is None:
            plt.show()
        else:
            fig.savefig(output, bbox_inches="tight")
            plt.close(fig)


def _build_layout(key: str, type: str) -> dict:
    """Load metadata and compute the common layout structures used by all plot functions.

    Returns a dict with keys: meta, str_pos, hpges, string_groups, strings,
    string_shade_map, sorted_cols, periods, period_colour_map, period_groups,
    usab_map, psd_map.
    """
    is_period_key = bool(re.match(r"^p\d+$", key))
    runinfo = Props.read_from("runinfo.yaml")
    runlists = Props.read_from("runlists.yaml") if not is_period_key else None

    meta = LegendMetadata()

    period_max: dict[str, int] = {}
    if is_period_key:
        runs_in_period = [r for r in runinfo[key] if type in runinfo[key][r]]
        if runs_in_period:
            period_max[key] = max(int(r[1:]) for r in runs_in_period)
    else:
        for period, runs_spec in runlists[key][type].items():
            runs = expand_run_list(runs_spec)
            if runs:
                period_max[period] = max(
                    period_max.get(period, 0), *(int(r[1:]) for r in runs)
                )

    sorted_cols = sorted(
        [(p, f"r{r:03d}") for p, mx in period_max.items() for r in range(mx + 1)],
        key=col_sort_key,
    )
    periods = sorted({p for p, _ in sorted_cols}, key=lambda p: int(p[1:]))
    period_colour_map = dict(
        zip(periods, cmap_hex(PERIOD_CMAP, max(len(periods), 1)), strict=False)
    )
    period_groups: dict[str, list] = {}
    for period, run in sorted_cols:
        period_groups.setdefault(period, []).append((period, run))

    # Build str_pos as a union across periods; the channelmap is stable within a period
    # so we use the first valid run of each period to get its layout.
    str_pos = {}
    for period in periods:
        period_info = runinfo.get(period, {})
        first_run = min(
            (r for r in period_info if type in period_info[r]),
            key=lambda r: int(r[1:]),
            default=None,
        )
        if first_run is None:
            continue
        timestamp = runinfo[period][first_run][type]["start_key"]
        period_chmap = meta.channelmap(timestamp)
        for ged, item in period_chmap.items():
            if item["system"] == "geds" and ged not in str_pos:
                str_pos[ged] = {
                    "string": item["location"]["string"],
                    "position": item["location"]["position"],
                }

    hpges = sorted(
        str_pos, key=lambda d: (str_pos[d]["string"], str_pos[d]["position"])
    )
    string_groups: dict[int, list] = {}
    for hpge in hpges:
        string_groups.setdefault(str_pos[hpge]["string"], []).append(hpge)
    strings = sorted(string_groups)
    string_shade_map = {s: STRING_CELL_SHADES[i % 2] for i, s in enumerate(strings)}

    usab_map = {}
    psd_map = {}
    for period, run in sorted_cols:
        if run not in runinfo.get(period, {}) or type not in runinfo[period][run]:
            continue
        timestamp = runinfo[period][run][type]["start_key"]
        run_chmap = meta.channelmap(timestamp)
        for hpge, item in run_chmap.items():
            if item["system"] != "geds":
                continue
            analysis = item.get("analysis")
            if not analysis:
                continue
            usability = analysis.get("usability")
            if usability is not None:
                usab_map[(period, run, hpge)] = usability
            if "psd" in analysis:
                psd_map[(period, run, hpge)] = analysis["psd"]

    return {
        "meta": meta,
        "str_pos": str_pos,
        "hpges": hpges,
        "string_groups": string_groups,
        "strings": strings,
        "string_shade_map": string_shade_map,
        "sorted_cols": sorted_cols,
        "periods": periods,
        "period_colour_map": period_colour_map,
        "period_groups": period_groups,
        "usab_map": usab_map,
        "psd_map": psd_map,
    }


def _render(layout: dict, hpge_maps: dict, cell_colours, output: str | None) -> None:
    """Render the grid to xlsx, pdf, or an interactive plot."""
    hpges = layout["hpges"]
    strings = layout["strings"]
    string_groups = layout["string_groups"]
    string_shade_map = layout["string_shade_map"]
    sorted_cols = layout["sorted_cols"]
    periods = layout["periods"]
    period_colour_map = layout["period_colour_map"]
    period_groups = layout["period_groups"]

    if output is not None and not output.endswith((".xlsx", ".pdf")):
        msg = f"Unsupported output format: {output!r}. Use '.xlsx' or '.pdf'."
        raise ValueError(msg)

    # ── xlsx ──────────────────────────────────────────────────────────────────

    if output is not None and output.endswith(".xlsx"):
        COL_OFFSET = 3
        GRP_ROW, HEADER_ROW, DATA_START = 1, 2, 3

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Groupings"

        ws.row_dimensions[GRP_ROW].height = 18
        col_cursor = COL_OFFSET
        for period in periods:
            runs = period_groups.get(period, [])
            if not runs:
                continue
            sc = col_cursor
            sl, el = get_column_letter(sc), get_column_letter(sc + len(runs) - 1)
            ws.merge_cells(f"{sl}{GRP_ROW}:{el}{GRP_ROW}")
            c = ws[f"{sl}{GRP_ROW}"]
            c.value = f"{period}  ({runs[0][1]} - {runs[-1][1]})"
            c.fill = xl_fill(period_colour_map[period])
            c.font = Font(bold=True, name=FONT, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            col_cursor += len(runs)

        ws.row_dimensions[HEADER_ROW].height = 70
        hdr_dark = xl_fill("2F4F7F")
        for col, label in enumerate(
            ["String", "Detector"] + [f"{p}_{r}" for p, r in sorted_cols], 1
        ):
            c = ws.cell(HEADER_ROW, col, label)
            c.fill = hdr_dark
            c.font = Font(bold=True, name=FONT, size=9, color="FFFFFF")
            c.alignment = (
                Alignment(horizontal="center", vertical="center", wrap_text=True)
                if col <= 2
                else Alignment(horizontal="center", vertical="bottom", text_rotation=90)
            )

        p_first_col = {}
        col_cursor = COL_OFFSET
        for period in periods:
            if period_groups.get(period):
                p_first_col[period] = col_cursor
                col_cursor += len(period_groups[period])
        last_col = COL_OFFSET + len(sorted_cols) - 1

        current_row = DATA_START
        for string_num in strings:
            string_hpges = string_groups[string_num]
            bg = string_shade_map[string_num]
            string_start = current_row

            for hpge_idx, hpge in enumerate(string_hpges):
                row = current_row
                is_first = hpge_idx == 0
                is_last = hpge_idx == len(string_hpges) - 1
                top = THICK if is_first else NORMAL
                bottom = THICK if is_last else None

                ca = ws.cell(row, 1)
                ca.fill = xl_fill(bg)
                ca.border = Border(
                    left=THICK,
                    right=NORMAL,
                    top=THICK if is_first else None,
                    bottom=THICK if is_last else None,
                )

                cb = ws.cell(row, 2, hpge)
                cb.fill = xl_fill(bg)
                cb.font = Font(bold=True, name=FONT, size=9)
                cb.alignment = Alignment(horizontal="center", vertical="center")
                cb.border = Border(left=NORMAL, right=THICK, top=top, bottom=bottom)

                for run_idx, (period, run) in enumerate(sorted_cols):
                    col = COL_OFFSET + run_idx
                    cell = ws.cell(row, col)
                    cell.border = Border(
                        left=THICK if col in p_first_col.values() else HAIR,
                        right=THICK if col == last_col else None,
                        top=top,
                        bottom=bottom,
                    )
                    fill_hex, lbl = cell_colours(hpge, period, run, hpge_maps[hpge])
                    cell.fill = xl_fill(fill_hex)
                    if lbl:
                        cell.value = lbl
                        cell.font = Font(name=FONT, size=8)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

                current_row += 1

            string_end = current_row - 1
            ws.merge_cells(f"A{string_start}:A{string_end}")
            c = ws.cell(string_start, 1)
            c.value = f"String {string_num}"
            c.font = Font(bold=True, name=FONT, size=10)
            c.fill = xl_fill(bg)
            c.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        ws.column_dimensions["A"].width = 7
        ws.column_dimensions["B"].width = 10
        for i in range(len(sorted_cols)):
            ws.column_dimensions[get_column_letter(COL_OFFSET + i)].width = 3.8
        for row in range(DATA_START, current_row):
            ws.row_dimensions[row].height = 15
        ws.freeze_panes = f"{get_column_letter(COL_OFFSET)}{DATA_START}"

        wb.save(output)

    # ── pdf / interactive ─────────────────────────────────────────────────────

    if output is None or output.endswith(".pdf"):
        n_det = len(hpges)
        n_cols = len(sorted_cols)

        label_grid = [[""] * n_cols for _ in range(n_det)]
        color_grid = np.ones((n_det, n_cols, 4))
        for row_i, det in enumerate(hpges):
            for col_i, (period, run) in enumerate(sorted_cols):
                fill_hex, lbl = cell_colours(det, period, run, hpge_maps[det])
                color_grid[row_i, col_i] = (*hex_to_rgb01(fill_hex), 1.0)
                label_grid[row_i][col_i] = lbl

        row_cursor = 0
        string_row_start = {}
        string_boundaries = []
        for string_num in strings:
            string_row_start[string_num] = row_cursor
            row_cursor += len(string_groups[string_num])
            string_boundaries.append(row_cursor)
        string_boundaries = string_boundaries[:-1]

        col_cursor = 0
        period_col_start = {}
        period_boundaries = []
        for period in periods:
            runs = period_groups.get(period, [])
            if not runs:
                continue
            period_col_start[period] = col_cursor
            col_cursor += len(runs)
            period_boundaries.append(col_cursor)
        period_boundaries = period_boundaries[:-1]

        STR_W, DET_W = 2.0, 3.5
        RUN_LBL_H, PER_BAR_H = 1.0, 0.8
        HDR_H = RUN_LBL_H + PER_BAR_H + 0.2
        x_str0, x_det0 = -(STR_W + DET_W), -DET_W
        cell_in = 0.28

        fig, ax = plt.subplots(
            figsize=(
                (STR_W + DET_W + n_cols) * cell_in + 0.2,
                (HDR_H + n_det) * cell_in + 0.3,
            )
        )
        ax.set_xlim(x_str0 - 0.05, n_cols + 0.05)
        ax.set_ylim(n_det, -HDR_H)
        ax.axis("off")

        for string_num in strings:
            start = string_row_start[string_num]
            size = len(string_groups[string_num])
            ax.add_patch(
                mpatches.Rectangle(
                    (x_str0, start),
                    n_cols - x_str0,
                    size,
                    facecolor=hex_to_rgb01(string_shade_map[string_num]),
                    edgecolor="none",
                    zorder=0,
                )
            )

        for row_i in range(n_det):
            for col_i in range(n_cols):
                ax.add_patch(
                    mpatches.Rectangle(
                        (col_i, row_i),
                        1,
                        1,
                        facecolor=color_grid[row_i, col_i],
                        edgecolor="#CCCCCC",
                        linewidth=0.3,
                        zorder=1,
                    )
                )
                lbl = label_grid[row_i][col_i]
                if lbl:
                    ax.text(
                        col_i + 0.5,
                        row_i + 0.5,
                        lbl,
                        ha="center",
                        va="center",
                        fontsize=5.5,
                        color="#333333",
                        zorder=2,
                    )

        for row_i, det in enumerate(hpges):
            ax.add_patch(
                mpatches.Rectangle(
                    (x_det0, row_i),
                    DET_W,
                    1,
                    facecolor="none",
                    edgecolor="#AAAAAA",
                    linewidth=0.3,
                    zorder=1,
                )
            )
            ax.text(
                x_det0 + DET_W / 2,
                row_i + 0.5,
                det,
                ha="center",
                va="center",
                fontsize=6,
                fontweight="bold",
                fontfamily="monospace",
                zorder=2,
            )

        for string_num in strings:
            start = string_row_start[string_num]
            size = len(string_groups[string_num])
            ax.add_patch(
                mpatches.Rectangle(
                    (x_str0, start),
                    STR_W,
                    size,
                    facecolor="none",
                    edgecolor="#555555",
                    linewidth=0.8,
                    zorder=3,
                )
            )
            ax.text(
                x_str0 + STR_W / 2,
                start + size / 2,
                f"Str\n{string_num}",
                ha="center",
                va="center",
                fontsize=6.5,
                fontweight="bold",
                zorder=4,
            )

        for sb in string_boundaries:
            ax.plot(
                [x_str0, n_cols], [sb, sb], color="#333333", linewidth=1.2, zorder=5
            )
        for pb in period_boundaries:
            ax.plot([pb, pb], [0, n_det], color="#333333", linewidth=1.2, zorder=5)

        bar_y0 = -(RUN_LBL_H + PER_BAR_H)
        bar_y_mid = bar_y0 + PER_BAR_H / 2
        for col_i, (_, run) in enumerate(sorted_cols):
            ax.text(
                col_i + 0.5,
                0,
                run,
                ha="center",
                va="bottom",
                fontsize=5.5,
                rotation=90,
                color="#222222",
                zorder=4,
            )

        for period in periods:
            runs = period_groups.get(period, [])
            if not runs:
                continue
            sc = period_col_start[period]
            ax.add_patch(
                mpatches.Rectangle(
                    (sc, bar_y0),
                    len(runs),
                    PER_BAR_H,
                    facecolor=hex_to_rgb01(period_colour_map[period]),
                    edgecolor="#555555",
                    linewidth=0.8,
                    zorder=3,
                )
            )
            ax.text(
                sc + len(runs) / 2,
                bar_y_mid,
                f"{period}  ({runs[0][1]}-{runs[-1][1]})",
                ha="center",
                va="center",
                fontsize=7,
                fontweight="bold",
                zorder=4,
            )

        for x0, w, label in [(x_str0, STR_W, "String"), (x_det0, DET_W, "Detector")]:
            ax.add_patch(
                mpatches.Rectangle(
                    (x0, bar_y0),
                    w,
                    PER_BAR_H,
                    facecolor=hex_to_rgb01("2F4F7F"),
                    edgecolor="#333333",
                    linewidth=0.8,
                    zorder=3,
                )
            )
            ax.text(
                x0 + w / 2,
                bar_y_mid,
                label,
                ha="center",
                va="center",
                fontsize=7,
                fontweight="bold",
                color="white",
                zorder=4,
            )

        plt.tight_layout(pad=0.2)
        if output is None:
            plt.show()
        else:
            fig.savefig(output, bbox_inches="tight")
            plt.close(fig)
